"""ORAD EM Controls export template.

Produces an Excel report with 20 columns:
  evaluation_date, control_id, level, parent_control_id,
  last_model_run_timestamp, last_model_run_parent_timestamp,
  7 W-criteria yes/no (inherited from parent for Level 2),
  7 operational criteria yes/no (control's own for Level 2).
"""

from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, case, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from server.pipelines.controls.schema import (
    ai_controls_model_enrichment,
    src_controls_rel_parent,
    src_controls_ver_control,
)

from ..base import ExportTemplate
from ..registry import register

# Column names reused from model_runners/run_enrichment_mock.py
W_CRITERIA_YES_NO = [
    "what_yes_no", "where_yes_no", "who_yes_no", "when_yes_no",
    "why_yes_no", "what_why_yes_no", "risk_theme_yes_no",
]
W_CRITERIA_DETAILS = [
    "what_details", "where_details", "who_details", "when_details",
    "why_details", "what_why_details", "risk_theme_details",
]
OP_CRITERIA_YES_NO = [
    "frequency_yes_no", "preventative_detective_yes_no",
    "automation_level_yes_no", "followup_yes_no",
    "escalation_yes_no", "evidence_yes_no", "abbreviations_yes_no",
]
OP_CRITERIA_DETAILS = [
    "frequency_details", "preventative_detective_details",
    "automation_level_details", "followup_details",
    "escalation_details", "evidence_details", "abbreviations_details",
]

# Paired: yes/no then details for each criterion
W_CRITERIA_PAIRED = [col for pair in zip(W_CRITERIA_YES_NO, W_CRITERIA_DETAILS) for col in pair]
OP_CRITERIA_PAIRED = [col for pair in zip(OP_CRITERIA_YES_NO, OP_CRITERIA_DETAILS) for col in pair]

EXPORT_COLUMNS = [
    "evaluation_date",
    "control_id",
    "level",
    "parent_control_id",
    "last_model_run_timestamp",
    "last_model_run_parent_timestamp",
    *W_CRITERIA_PAIRED,
    *OP_CRITERIA_PAIRED,
]


def _as_of(table, evaluation_date: datetime):
    """Temporal filter: row was current at evaluation_date."""
    return and_(
        table.c.tx_from <= evaluation_date,
        or_(table.c.tx_to.is_(None), table.c.tx_to > evaluation_date),
    )


@register
class OradEmControlsTemplate(ExportTemplate):
    TEMPLATE_NAME = "orad_em_controls"
    TEMPLATE_DESCRIPTION = (
        "Enrichment model W-criteria scores for all active key controls. "
        "Level 2 controls inherit the first 7 W-criteria from their parent."
    )

    async def query(self, db: AsyncSession, evaluation_date: datetime) -> List[Dict]:
        vc = src_controls_ver_control.alias("vc")
        enr = ai_controls_model_enrichment.alias("enr")

        # Build a self-contained subquery for parent info:
        #   child_control_id → parent_control_id, parent W-criteria, parent timestamp
        # This resolves: rel_parent → parent enrichment in one subquery,
        # avoiding chained outerjoin issues.
        rel = src_controls_rel_parent.alias("rel")
        p_enr = ai_controls_model_enrichment.alias("p_enr")

        # Parent subquery includes both yes/no and details for W-criteria
        w_all = W_CRITERIA_YES_NO + W_CRITERIA_DETAILS
        parent_sub = (
            select(
                rel.c.child_control_id,
                rel.c.parent_control_id,
                p_enr.c.model_run_timestamp.label("parent_model_run_timestamp"),
                *[p_enr.c[col].label(f"p_{col}") for col in w_all],
            )
            .select_from(
                rel.join(p_enr, and_(
                    p_enr.c.ref_control_id == rel.c.parent_control_id,
                    _as_of(p_enr, evaluation_date),
                ))
            )
            .where(_as_of(rel, evaluation_date))
            .subquery("parent_info")
        )

        is_level2 = vc.c.hierarchy_level == "Level 2"

        # W-criteria (yes/no + details): from parent for L2, own for L1
        w_columns = []
        for col_name in w_all:
            w_columns.append(
                case(
                    (is_level2, parent_sub.c[f"p_{col_name}"]),
                    else_=enr.c[col_name],
                ).label(col_name)
            )

        # Operational criteria (yes/no + details): always from own enrichment
        op_all = OP_CRITERIA_YES_NO + OP_CRITERIA_DETAILS
        op_columns = [enr.c[col_name].label(col_name) for col_name in op_all]

        stmt = (
            select(
                vc.c.ref_control_id.label("control_id"),
                vc.c.hierarchy_level.label("level"),
                case(
                    (is_level2, parent_sub.c.parent_control_id),
                    else_=None,
                ).label("parent_control_id"),
                enr.c.model_run_timestamp.label("last_model_run_timestamp"),
                case(
                    (is_level2, parent_sub.c.parent_model_run_timestamp),
                    else_=None,
                ).label("last_model_run_parent_timestamp"),
                *w_columns,
                *op_columns,
            )
            .select_from(
                vc
                .join(enr, and_(
                    enr.c.ref_control_id == vc.c.ref_control_id,
                    _as_of(enr, evaluation_date),
                ))
                .outerjoin(parent_sub,
                    parent_sub.c.child_control_id == vc.c.ref_control_id,
                )
            )
            .where(
                _as_of(vc, evaluation_date),
                vc.c.control_status == "Active",
                vc.c.key_control.is_(True),
            )
            .order_by(vc.c.ref_control_id)
        )

        result = await db.execute(stmt)
        rows = [dict(row._mapping) for row in result.fetchall()]

        if not rows:
            await self._diagnose_empty(db, evaluation_date)

        self._validate(rows)
        return rows

    async def _diagnose_empty(
        self, db: AsyncSession, evaluation_date: datetime
    ) -> None:
        """When query returns 0 rows, figure out why and raise a clear error."""
        vc = src_controls_ver_control.alias("vc_diag")
        enr = ai_controls_model_enrichment.alias("enr_diag")

        # Check 1: any controls at all as-of this date?
        ctrl_count = await db.scalar(
            select(func.count())
            .select_from(vc)
            .where(_as_of(vc, evaluation_date))
        )
        if ctrl_count == 0:
            raise ValueError(
                f"No controls found as of {evaluation_date.date()}. "
                "The controls table may be empty or no data existed at this date."
            )

        # Check 2: any active key controls?
        active_key_count = await db.scalar(
            select(func.count())
            .select_from(vc)
            .where(
                _as_of(vc, evaluation_date),
                vc.c.control_status == "Active",
                vc.c.key_control.is_(True),
            )
        )
        if active_key_count == 0:
            raise ValueError(
                f"Found {ctrl_count} controls as of {evaluation_date.date()}, "
                "but none are Active with key_control=True."
            )

        # Check 3: any enrichment data?
        enr_count = await db.scalar(
            select(func.count())
            .select_from(enr)
            .where(_as_of(enr, evaluation_date))
        )
        if enr_count == 0:
            raise ValueError(
                f"Found {active_key_count} active key controls as of "
                f"{evaluation_date.date()}, but the enrichment table is empty "
                "at this date. Run enrichment model before exporting."
            )

        # Fallback: controls and enrichment exist but don't join
        raise ValueError(
            f"Found {active_key_count} active key controls and "
            f"{enr_count} enrichment records as of {evaluation_date.date()}, "
            "but no matching control-enrichment pairs. "
            "Check that enrichment has been ingested for these controls."
        )

    def _validate(self, rows: List[Dict]) -> None:
        """Ensure no yes/no field is NULL. Raise with offending control IDs."""
        errors: List[str] = []
        for row in rows:
            level = row.get("level", "")
            is_l2 = str(level).strip().lower() == "level 2"

            # For L1: check first 7 W-criteria (own)
            # For L2: check first 7 W-criteria (from parent) + second 7 (own)
            cols_to_check = list(W_CRITERIA_YES_NO)
            if is_l2:
                cols_to_check += OP_CRITERIA_YES_NO

            missing = [c for c in cols_to_check if row.get(c) is None]
            if missing:
                # Distinguish missing parent enrichment from missing own fields
                parent_missing = [c for c in missing if c in W_CRITERIA_YES_NO] if is_l2 else []
                own_missing = [c for c in missing if c in OP_CRITERIA_YES_NO] if is_l2 else missing

                parts = []
                if parent_missing:
                    parent_id = row.get("parent_control_id", "unknown")
                    parts.append(
                        f"parent {parent_id} enrichment missing for {parent_missing}"
                    )
                if own_missing:
                    parts.append(f"own enrichment missing for {own_missing}")
                detail = ", ".join(parts) if parts else f"NULL in {missing}"

                errors.append(
                    f"control_id={row['control_id']} (level={level}): {detail}"
                )

        if errors:
            detail = "; ".join(errors[:20])  # cap error list
            raise ValueError(
                f"Export validation failed — {len(errors)} control(s) have NULL "
                f"yes/no fields: {detail}"
            )

    def build_workbook(self, rows: List[Dict], evaluation_date: datetime) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "ORAD EM Controls"

        # Header row
        bold = Font(bold=True)
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = bold

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=evaluation_date.date())
            for col_idx, col_name in enumerate(EXPORT_COLUMNS[1:], start=2):
                value = row.get(col_name)
                # Strip tzinfo — Excel does not support tz-aware datetimes
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths
        for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
            max_len = len(str(EXPORT_COLUMNS[col_idx - 1]))
            for row_idx in range(2, min(len(rows) + 2, 102)):  # sample first 100 rows
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        return wb
